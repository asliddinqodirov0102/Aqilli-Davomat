from flask import (
    Flask, render_template, request, jsonify,
    session, redirect, url_for
)
from functools import wraps
import math
import sqlite3
import os
import uuid
import base64
import io
from datetime import datetime, timezone
from werkzeug.security import generate_password_hash, check_password_hash

try:
    import qrcode
    from qrcode.image.pure import PyPNGImage
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

app = Flask(__name__)

# Secret key for signing sessions – override via environment variable in production
app.secret_key = os.environ.get("SECRET_KEY", "default_secret_key")

DATABASE = "attendance.db"


# ── Database helpers ────────────────────────────────────────────────────────

def get_db():
    """Create a database connection."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


# Session expires after this many seconds (5 minutes)
SESSION_TTL = 300

# Default classroom location (Tashkent) and allowed radius in metres
CLASSROOM_LAT = 41.2995
CLASSROOM_LON = 69.2401
MAX_DISTANCE_M = 50


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Return the great-circle distance in metres between two GPS points."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi  = math.radians(lat2 - lat1)
    dlam  = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def init_db():
    """Initialize the database and create tables if they don't exist."""
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT    NOT NULL,
            role      TEXT    NOT NULL,
            username  TEXT    NOT NULL UNIQUE,
            password  TEXT    NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            token       TEXT    NOT NULL UNIQUE,
            teacher_id    INTEGER NOT NULL,
            created_at    TEXT    NOT NULL,
            is_active     INTEGER NOT NULL DEFAULT 1,
            classroom_lat REAL,
            classroom_lon REAL,
            subject       TEXT,
            FOREIGN KEY (teacher_id) REFERENCES users(id)
        )
    """)
    # Migration: add columns to existing databases
    for col_sql in [
        "ALTER TABLE attendance_sessions ADD COLUMN classroom_lat REAL",
        "ALTER TABLE attendance_sessions ADD COLUMN classroom_lon REAL",
        "ALTER TABLE attendance_sessions ADD COLUMN subject TEXT",
    ]:
        try:
            conn.execute(col_sql)
        except Exception:
            pass  # column already exists
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_records (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id    INTEGER NOT NULL,
            session_token TEXT    NOT NULL,
            marked_at     TEXT    NOT NULL,
            latitude      REAL,
            longitude     REAL,
            distance_m    REAL,
            status        TEXT    NOT NULL DEFAULT 'present',
            FOREIGN KEY (student_id)    REFERENCES users(id),
            FOREIGN KEY (session_token) REFERENCES attendance_sessions(token),
            UNIQUE (student_id, session_token)
        )
    """)
    conn.commit()
    conn.close()


def _make_qr_b64(data: str) -> str:
    """Generate a QR code for *data* and return it as a base64-encoded PNG string."""
    if not QR_AVAILABLE:
        return ""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=20,
        border=6,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#000000", back_color="#ffffff")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


# ── Auth decorator ──────────────────────────────────────────────────────────

def login_required(f):
    """Redirect unauthenticated visitors to /login, preserving the target URL."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            # Save the full requested URL so we can return after login
            return redirect(url_for("login") + "?next=" + request.url)
        return f(*args, **kwargs)
    return decorated


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return render_template("register.html")


# ── Registration ─────────────────────────────────────────────────────────────

@app.route("/register", methods=["POST"])
def register():
    """Handle user registration via JSON body or form data."""
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form

    full_name = (data.get("full_name") or "").strip()
    role      = (data.get("role")      or "").strip()
    username  = (data.get("username")  or "").strip()
    password  = (data.get("password")  or "").strip()

    errors = {}
    if not full_name:
        errors["full_name"] = "To'liq ism majburiy."
    if role not in ("student", "teacher"):
        errors["role"] = "Rol noto'g'ri."
    if not username:
        errors["username"] = "Foydalanuvchi nomi majburiy."
    elif len(username) < 3:
        errors["username"] = "Foydalanuvchi nomi kamida 3 ta belgidan iborat bo'lishi kerak."
    if not password:
        errors["password"] = "Parol majburiy."
    elif len(password) < 6:
        errors["password"] = "Parol kamida 6 ta belgidan iborat bo'lishi kerak."

    if errors:
        return jsonify({"success": False, "errors": errors}), 422

    hashed_pw = generate_password_hash(password)

    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO users (full_name, role, username, password) VALUES (?, ?, ?, ?)",
            (full_name, role, username, hashed_pw),
        )
        conn.commit()
        conn.close()
    except sqlite3.IntegrityError:
        return jsonify({
            "success": False,
            "errors": {"username": "Bu foydalanuvchi nomi allaqachon band. Boshqa nom tanlang."}
        }), 409

    return jsonify({
        "success": True,
        "message": "Ro'yxatdan muvaffaqiyatli o'tdingiz!"
    }), 201


# ── Login ────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login_post():
    """Authenticate the user and store data in session."""
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form

    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()

    errors = {}
    if not username:
        errors["username"] = "Foydalanuvchi nomi majburiy."
    if not password:
        errors["password"] = "Parol majburiy."
    if errors:
        return jsonify({"success": False, "errors": errors}), 422

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ?", (username,)
    ).fetchone()
    conn.close()

    if user is None or not check_password_hash(user["password"], password):
        return jsonify({
            "success": False,
            "errors": {"password": "Foydalanuvchi nomi yoki parol noto'g'ri."}
        }), 401

    # Store minimal info in the signed session cookie
    session["user_id"]   = user["id"]
    session["username"]  = user["username"]
    session["full_name"] = user["full_name"]
    session["role"]      = user["role"]

    # Honour ?next= (e.g. /checkin/<token> saved by login_required)
    next_url = request.args.get("next") or request.get_json(silent=True, force=True) and None
    if not next_url:
        data_again = request.get_json(silent=True) or {}
        next_url = data_again.get("next", "")
    if next_url and next_url.startswith("http"):
        # Absolute URL — only allow same host for security
        from urllib.parse import urlparse
        parsed = urlparse(next_url)
        if parsed.netloc == request.host:
            redirect_url = next_url
        else:
            redirect_url = url_for(
                "teacher_dashboard" if user["role"] == "teacher" else "student_dashboard"
            )
    elif next_url and next_url.startswith("/"):
        redirect_url = next_url
    else:
        redirect_url = url_for(
            "teacher_dashboard" if user["role"] == "teacher" else "student_dashboard"
        )
    return jsonify({"success": True, "redirect": redirect_url}), 200


# ── Generic dashboard redirect ────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    if session["role"] == "teacher":
        return redirect(url_for("teacher_dashboard"))
    return redirect(url_for("student_dashboard"))


# ── Teacher dashboard ─────────────────────────────────────────────────────────

@app.route("/teacher_dashboard")
@login_required
def teacher_dashboard():
    if session["role"] != "teacher":
        return redirect(url_for("student_dashboard"))
    return render_template("teacher_dashboard.html",
                           full_name=session["full_name"],
                           username=session["username"])


# ── Student dashboard ─────────────────────────────────────────────────────────

@app.route("/student_dashboard")
@login_required
def student_dashboard():
    if session["role"] != "student":
        return redirect(url_for("teacher_dashboard"))
    return render_template("student_dashboard.html",
                           full_name=session["full_name"],
                           username=session["username"])


# ── Logout ────────────────────────────────────────────────────────────────────

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── QR / Attendance Session APIs ─────────────────────────────────────────────

@app.route("/api/start_session", methods=["POST"])
@login_required
def start_session():
    """Create a new attendance session, expire any previous active ones, return QR."""
    if session["role"] != "teacher":
        return jsonify({"error": "Ruxsat yo'q."}), 403

    if not QR_AVAILABLE:
        return jsonify({"error": "qrcode kutubxonasi o'rnatilmagan. 'pip install qrcode[pil]' buyrug'ini ishga tushiring."}), 500

    data = request.get_json(silent=True) or {}
    subject = (data.get("subject") or "").strip()
    if not subject:
        return jsonify({"error": "Fan nomi kiritilmagan."}), 422

    teacher_id = session["user_id"]
    token = uuid.uuid4().hex
    now   = datetime.now(timezone.utc).isoformat()

    conn = get_db()
    conn.execute(
        "UPDATE attendance_sessions SET is_active = 0 WHERE teacher_id = ? AND is_active = 1",
        (teacher_id,)
    )
    conn.execute(
        "INSERT INTO attendance_sessions (token, teacher_id, created_at, is_active, subject) VALUES (?, ?, ?, 1, ?)",
        (token, teacher_id, now, subject)
    )
    conn.commit()
    conn.close()

    checkin_url = request.host_url.rstrip("/") + f"/checkin/{token}"
    qr_b64 = _make_qr_b64(checkin_url)

    return jsonify({
        "success":    True,
        "token":      token,
        "qr_b64":     qr_b64,
        "created_at": now,
        "expires_in": SESSION_TTL,
        "checkin_url": checkin_url,
        "subject":    subject,
    }), 201


@app.route("/api/session_status/<token>", methods=["GET"])
@login_required
def session_status(token):
    """Return whether the token is still valid and how many students checked in."""
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM attendance_sessions WHERE token = ?", (token,)
    ).fetchone()

    if row is None:
        conn.close()
        return jsonify({"valid": False, "reason": "Token topilmadi."}), 404

    created_at = datetime.fromisoformat(row["created_at"])
    now        = datetime.now(timezone.utc)
    age        = (now - created_at).total_seconds()
    is_active  = bool(row["is_active"])

    if not is_active or age > SESSION_TTL:
        if is_active and age > SESSION_TTL:
            conn.execute(
                "UPDATE attendance_sessions SET is_active = 0 WHERE token = ?", (token,)
            )
            conn.commit()
        conn.close()
        return jsonify({"valid": False, "reason": "Sessiya muddati tugagan.", "age": int(age)}), 200

    # Fetch checked-in students list
    records = conn.execute(
        """
        SELECT u.full_name, u.username, r.marked_at, r.distance_m
        FROM attendance_records r
        JOIN users u ON u.id = r.student_id
        WHERE r.session_token = ?
        ORDER BY r.marked_at
        """, (token,)
    ).fetchall()

    conn.close()
    remaining = max(0, SESSION_TTL - int(age))
    return jsonify({
        "valid":         True,
        "token":         token,
        "checkin_count": len(records),
        "age":           int(age),
        "remaining":     remaining,
        "students":      [{"full_name": r["full_name"], "username": r["username"],
                           "marked_at": r["marked_at"], "distance_m": r["distance_m"]} for r in records],
    }), 200


# ── Set classroom location ─────────────────────────────────────────────────────

@app.route("/api/set_classroom_location", methods=["POST"])
@login_required
def set_classroom_location():
    """Save the teacher's current GPS as the classroom anchor for this session."""
    if session["role"] != "teacher":
        return jsonify({"error": "Ruxsat yo'q."}), 403
    data = request.get_json()
    try:
        lat = float(data.get("latitude",  0))
        lon = float(data.get("longitude", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "GPS koordinatalari noto'g'ri."}), 422
    token = (data.get("token") or "").strip()
    if not token:
        return jsonify({"error": "Avval sessiya yarating."}), 422
    conn = get_db()
    result = conn.execute(
        "UPDATE attendance_sessions SET classroom_lat=?, classroom_lon=? "
        "WHERE token=? AND teacher_id=?",
        (lat, lon, token, session["user_id"])
    )
    conn.commit()
    conn.close()
    if result.rowcount == 0:
        return jsonify({"error": "Faol sessiya topilmadi."}), 404
    return jsonify({
        "success":   True,
        "message":   "Sinf joylashuvi saqlandi!",
        "latitude":  lat,
        "longitude": lon,
    }), 200


# ── Student check-in routes ──────────────────────────────────────────────────

@app.route("/checkin/<token>")
@login_required
def checkin_redirect(token):
    """QR scan lands here; redirect student dashboard with token pre-filled."""
    if session["role"] != "student":
        return redirect(url_for("teacher_dashboard"))
    return redirect(url_for("student_dashboard") + f"?token={token}")


@app.route("/mark_attendance", methods=["POST"])
@login_required
def mark_attendance():
    """Verify token + GPS distance, then record attendance."""
    if session["role"] != "student":
        return jsonify({"success": False, "error": "Faqat talabalar uchun."}), 403

    data      = request.get_json()
    token     = (data.get("token")     or "").strip()
    try:
        lat = float(data.get("latitude",  0))
        lon = float(data.get("longitude", 0))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "GPS koordinatalari noto'g'ri."}), 422

    if not token:
        return jsonify({"success": False, "error": "Token mavjud emas."}), 422

    conn = get_db()

    # ── Check 1: token validity ────────────────────────────────────────────
    row = conn.execute(
        "SELECT * FROM attendance_sessions WHERE token = ?", (token,)
    ).fetchone()

    if row is None:
        conn.close()
        return jsonify({"success": False, "error": "QR kod topilmadi. Yangi QR ko'rsating."}), 404

    created_at = datetime.fromisoformat(row["created_at"])
    age = (datetime.now(timezone.utc) - created_at).total_seconds()

    if not row["is_active"] or age > SESSION_TTL:
        conn.close()
        return jsonify({"success": False, "error": "QR kod muddati tugagan. O'qituvchidan yangi QR so'rang."}), 410

    # ── Check 2: GPS distance (use session-specific classroom if set) ──────────
    class_lat = row["classroom_lat"] if row["classroom_lat"] is not None else CLASSROOM_LAT
    class_lon = row["classroom_lon"] if row["classroom_lon"] is not None else CLASSROOM_LON
    distance = haversine_m(lat, lon, class_lat, class_lon)
    if distance > MAX_DISTANCE_M:
        conn.close()
        return jsonify({
            "success":  False,
            "error":    f"Siz sinf xonasidan juda uzoqdasiz ({distance:.0f}m). Ruxsat etilgan masofa: {MAX_DISTANCE_M}m.",
            "distance": round(distance, 1),
        }), 403

    # ── Record attendance ──────────────────────────────────────────────────
    student_id = session["user_id"]
    now        = datetime.now(timezone.utc).isoformat()

    try:
        conn.execute(
            """
            INSERT INTO attendance_records
                (student_id, session_token, marked_at, latitude, longitude, distance_m, status)
            VALUES (?, ?, ?, ?, ?, ?, 'present')
            """,
            (student_id, token, now, lat, lon, round(distance, 2))
        )
        conn.commit()
    except Exception as e:
        conn.close()
        if "UNIQUE" in str(e):
            return jsonify({"success": False, "error": "Siz bu dars uchun allaqachon davomat qildingiz."}), 409
        return jsonify({"success": False, "error": "Xatolik yuz berdi. Qayta urinib ko'ring."}), 500

    conn.close()
    return jsonify({
        "success":   True,
        "message":   "Davomat muvaffaqiyatli qayd etildi!",
        "distance":  round(distance, 1),
        "marked_at": now,
    }), 201


# ── Excel Export ─────────────────────────────────────────────────────────────

@app.route("/api/export_session/<token>")
@login_required
def export_session(token):
    """Export attendance records for a session as a styled .xlsx file."""
    if session["role"] != "teacher":
        return jsonify({"error": "Ruxsat yo'q."}), 403

    if not XLSX_AVAILABLE:
        return jsonify({"error": "openpyxl o'rnatilmagan. 'pip install openpyxl' buyrug'ini ishga tushiring."}), 500

    conn = get_db()

    # Verify token belongs to this teacher
    sess_row = conn.execute(
        "SELECT * FROM attendance_sessions WHERE token = ? AND teacher_id = ?",
        (token, session["user_id"])
    ).fetchone()
    if sess_row is None:
        conn.close()
        return jsonify({"error": "Sessiya topilmadi."}), 404

    # Fetch attendance records joined with student info
    rows = conn.execute(
        """
        SELECT
            u.full_name,
            u.username,
            r.marked_at,
            r.distance_m,
            r.status
        FROM attendance_records r
        JOIN users u ON u.id = r.student_id
        WHERE r.session_token = ?
        ORDER BY r.marked_at
        """,
        (token,)
    ).fetchall()
    conn.close()

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"

    # Colour palette
    HDR_FILL   = PatternFill("solid", fgColor="1e3a2f")   # dark emerald
    ALT_FILL   = PatternFill("solid", fgColor="0f1e19")   # subtle alternate
    ACCENT     = PatternFill("solid", fgColor="059669")   # emerald-600
    thin_side  = Side(style="thin", color="2d4a3e")
    border     = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Title row ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    subject_name = sess_row["subject"] or "Noma'lum fan"
    title_cell.value     = f"Davomat hisoboti — {subject_name} ({token[:8]}…)"
    title_cell.font      = Font(bold=True, size=14, color="10B981", name="Calibri")
    title_cell.fill      = PatternFill("solid", fgColor="0a1512")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Meta rows ─────────────────────────────────────────────────────────
    meta_font = Font(color="6EE7B7", name="Calibri", size=10)
    ws["A2"] = "Fan:"
    ws["B2"] = sess_row["subject"] or "—"
    ws["A3"] = "Sessiya tokeni:"
    ws["B3"] = token
    ws["A4"] = "Eksport vaqti:"
    ws["B4"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A5"] = "Jami talabalar:"
    ws["B5"] = len(rows)
    for r in range(2, 6):
        ws.cell(r, 1).font = Font(bold=True, color="A7F3D0", name="Calibri", size=10)
        ws.cell(r, 2).font = meta_font
        for c in range(1, 6):
            ws.cell(r, c).fill = PatternFill("solid", fgColor="0d1f18")

    ws.row_dimensions[6].height = 8  # spacer

    # ── Header row ────────────────────────────────────────────────────────
    headers = ["#", "To'liq Ism", "Login (Username)", "Vaqt (UTC)", "Masofa (m)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(7, col)
        cell.value     = h
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[7].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, row in enumerate(rows, 1):
        xlsx_row = i + 7
        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="111f18")
        # Format timestamp: strip microseconds, make readable
        try:
            ts = datetime.fromisoformat(row["marked_at"]).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            ts = row["marked_at"]

        values = [i, row["full_name"], row["username"], ts, round(row["distance_m"] or 0, 1)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(xlsx_row, col)
            cell.value     = val
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "left",
                                       vertical="center")
            cell.font      = Font(color="D1FAE5" if col != 1 else "6EE7B7",
                                  name="Calibri", size=10)
        ws.row_dimensions[xlsx_row].height = 18

    # No-data notice
    if not rows:
        ws.cell(8, 1).value = "Hech kim davomat qilmagan."
        ws.cell(8, 1).font  = Font(color="6B7280", italic=True, name="Calibri")
        ws.merge_cells("A8:E8")

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [5, 28, 22, 22, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Stream to browser ────────────────────────────────────────────────
    from flask import Response
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"davomat_{token[:8]}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Entry point ───────────────────────────────────────────────────────────────

# Always initialise DB so Render's ephemeral container is ready on cold start
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"🚀  Server ishga tushmoqda: http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port)
